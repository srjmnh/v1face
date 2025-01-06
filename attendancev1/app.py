import os
import sys
import base64
import json
import io
from datetime import datetime
import logging

import cv2
import numpy as np
from PIL import Image, ImageEnhance
from flask import Flask, request, jsonify, render_template_string, send_file

# -----------------------------
# 1) AWS Rekognition Setup
# -----------------------------
import boto3

AWS_ACCESS_KEY_ID = os.getenv('AWS_ACCESS_KEY_ID')
AWS_SECRET_ACCESS_KEY = os.getenv('AWS_SECRET_ACCESS_KEY')
AWS_REGION = os.getenv('AWS_REGION', 'us-east-1')
COLLECTION_ID = "students"

rekognition_client = boto3.client(
    'rekognition',
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY,
    region_name=AWS_REGION
)

def create_collection_if_not_exists(collection_id):
    try:
        rekognition_client.create_collection(CollectionId=collection_id)
        print(f"Collection '{collection_id}' created.")
    except rekognition_client.exceptions.ResourceAlreadyExistsException:
        print(f"Collection '{collection_id}' already exists.")

create_collection_if_not_exists(COLLECTION_ID)

# -----------------------------
# 2) Firebase Firestore Setup
# -----------------------------
import firebase_admin
from firebase_admin import credentials, firestore

base64_cred_str = os.environ.get("FIREBASE_ADMIN_CREDENTIALS_BASE64")
if not base64_cred_str:
    raise ValueError("FIREBASE_ADMIN_CREDENTIALS_BASE64 not found in environment.")

decoded_cred_json = base64.b64decode(base64_cred_str)
cred_dict = json.loads(decoded_cred_json)
cred = credentials.Certificate(cred_dict)
firebase_admin.initialize_app(cred)
db = firestore.client()

# -----------------------------
# 3) Gemini Chatbot Setup
# -----------------------------
import google.generativeai as genai

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    raise ValueError("GEMINI_API_KEY environment variable not set.")

genai.configure(api_key=GEMINI_API_KEY)
# If you do NOT have access to "gemini-1.5-flash", switch to "models/chat-bison-001"
model = genai.GenerativeModel("models/gemini-1.5-flash")

# Chat memory
MAX_MEMORY = 20
conversation_memory = []

# A big system prompt describing the entire system
system_context = """You are Gemini, a somewhat witty (but polite) AI assistant.
Facial Recognition Attendance system features:

1) AWS Rekognition:
   - We keep a 'students' collection on startup.
   - /register indexes a face by (name + student_id).
   - /recognize detects faces in an uploaded image, logs attendance in Firestore if matched.

2) Attendance:
   - Firestore 'attendance' collection: { student_id, name, timestamp, subject_id, subject_name, status='PRESENT' }.
   - UI has tabs: Register, Recognize, Subjects, Attendance (Bootstrap + DataTables).
   - Attendance can filter, inline-edit, download/upload Excel.

3) Subjects:
   - We can add subjects to 'subjects' collection, referenced in recognition.

4) Multi-Face:
   - If multiple recognized faces, each is logged to attendance.

5) Chat:
   - You are the assistant, a bit humorous, guiding usage or code features.
"""

# Start the conversation with a system message
conversation_memory.append({"role": "system", "content": system_context})

# -----------------------------
# 4) Flask App
# -----------------------------
app = Flask(__name__)

# -----------------------------
# 5) Image Enhancement Function (Using OpenCV and Pillow)
# -----------------------------
def enhance_image(pil_image):
    """
    Enhance image quality to improve face detection in distant group photos.
    This includes increasing brightness and contrast.
    """
    # Convert PIL image to OpenCV format
    cv_image = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)

    # Increase brightness and contrast
    alpha = 1.2  # Contrast control (1.0-3.0)
    beta = 30    # Brightness control (0-100)
    enhanced_cv_image = cv2.convertScaleAbs(cv_image, alpha=alpha, beta=beta)

    # Convert back to PIL Image
    enhanced_pil_image = Image.fromarray(cv2.cvtColor(enhanced_cv_image, cv2.COLOR_BGR2RGB))

    return enhanced_pil_image

# -----------------------------
# 6) Single-Page HTML + Chat Widget
# -----------------------------
INDEX_HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <title>Facial Recognition Attendance + Gemini Chat</title>
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <!-- Bootstrap CSS -->
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet" />
  <!-- DataTables CSS -->
  <link rel="stylesheet" href="https://cdn.datatables.net/1.13.4/css/jquery.dataTables.min.css">
  <style>
    body { margin: 20px; }
    .nav-tabs .nav-link { color: #555; }
    .nav-tabs .nav-link.active { color: #000; font-weight: bold; }
    #attendanceTable td[contenteditable="true"] { background-color: #fcf8e3; }

    /* Chatbot Toggle Button */
    #chatbotToggle {
      position: fixed;
      bottom: 20px;
      right: 20px;
      z-index: 999;
      background-color: #0d6efd;
      color: #fff;
      border: none;
      border-radius: 50%;
      width: 50px;
      height: 50px;
      font-size: 22px;
      cursor: pointer;
      display: flex;
      align-items: center;
      justify-content: center;
    }
    #chatbotToggle:hover { background-color: #0b5ed7; }

    /* Chat Window */
    #chatbotWindow {
      position: fixed;
      bottom: 80px;
      right: 20px;
      width: 300px;
      max-height: 400px;
      border: 1px solid #ccc;
      border-radius: 10px;
      background-color: #fff;
      box-shadow: 0 0 10px rgba(0,0,0,0.2);
      display: none;
      flex-direction: column;
      z-index: 1000;
    }
    #chatHeader {
      background-color: #0d6efd;
      color: #fff;
      padding: 10px;
      border-radius: 10px 10px 0 0;
      font-weight: bold;
      display: flex;
      justify-content: space-between;
      align-items: center;
    }
    #chatHeader .close-btn {
      background: none;
      border: none;
      color: #fff;
      font-weight: bold;
      cursor: pointer;
      font-size: 16px;
    }
    #chatMessages {
      flex: 1;
      overflow-y: auto;
      padding: 10px;
      font-size: 14px;
    }
    .message { margin-bottom: 10px; }
    .message.user { text-align: right; }
    .message.assistant { text-align: left; color: #333; }
    #chatInputArea {
      display: flex; border-top: 1px solid #ddd;
    }
    #chatInput {
      flex: 1; padding: 8px; border: none; outline: none; font-size: 14px;
    }
    #chatSendBtn {
      background-color: #0d6efd; color: #fff;
      border: none; padding: 0 15px; cursor: pointer;
    }
    #chatSendBtn:hover { background-color: #0b5ed7; }
  </style>
</head>
<body class="container">

<h1 class="my-4">Facial Recognition Attendance + Gemini Chat</h1>

<!-- Nav Tabs -->
<ul class="nav nav-tabs" id="mainTabs" role="tablist">
  <li class="nav-item">
    <button class="nav-link active" id="register-tab" data-bs-toggle="tab" data-bs-target="#register" type="button" role="tab">
      Register
    </button>
  </li>
  <li class="nav-item">
    <button class="nav-link" id="recognize-tab" data-bs-toggle="tab" data-bs-target="#recognize" type="button" role="tab">
      Recognize
    </button>
  </li>
  <li class="nav-item">
    <button class="nav-link" id="subjects-tab" data-bs-toggle="tab" data-bs-target="#subjects" type="button" role="tab">
      Subjects
    </button>
  </li>
  <li class="nav-item">
    <button class="nav-link" id="attendance-tab" data-bs-toggle="tab" data-bs-target="#attendance" type="button" role="tab">
      Attendance
    </button>
  </li>
</ul>

<div class="tab-content" id="mainTabContent">
  <!-- REGISTER -->
  <div class="tab-pane fade show active mt-4" id="register" role="tabpanel" aria-labelledby="register-tab">
    <h3>Register a Face</h3>
    <label class="form-label">Name</label>
    <input type="text" id="reg_name" class="form-control" placeholder="Enter Name" />
    <label class="form-label">Student ID</label>
    <input type="text" id="reg_student_id" class="form-control" placeholder="Enter Student ID" />
    <label class="form-label">Image</label>
    <input type="file" id="reg_image" class="form-control" accept="image/*" />
    <button onclick="registerFace()" class="btn btn-primary mt-2">Register</button>
    <div id="register_result" class="alert alert-info mt-3" style="display:none;"></div>
  </div>

  <!-- RECOGNIZE -->
  <div class="tab-pane fade mt-4" id="recognize" role="tabpanel" aria-labelledby="recognize-tab">
    <h3>Recognize Faces</h3>
    <label class="form-label">Subject (optional)</label>
    <select id="rec_subject_select" class="form-control mb-2">
      <option value="">-- No Subject --</option>
    </select>
    <label class="form-label">Image</label>
    <input type="file" id="rec_image" class="form-control" accept="image/*" />
    <button onclick="recognizeFace()" class="btn btn-success mt-2">Recognize</button>
    <div id="recognize_result" class="alert alert-info mt-3" style="display:none;"></div>
  </div>

  <!-- SUBJECTS -->
  <div class="tab-pane fade mt-4" id="subjects" role="tabpanel" aria-labelledby="subjects-tab">
    <h3>Manage Subjects</h3>
    <label class="form-label">New Subject Name:</label>
    <input type="text" id="subject_name" class="form-control" placeholder="e.g. Mathematics" />
    <button onclick="addSubject()" class="btn btn-primary mt-2">Add Subject</button>
    <div id="subject_result" class="alert alert-info mt-3" style="display:none;"></div>
    <hr />
    <h5>Existing Subjects</h5>
    <ul id="subjects_list"></ul>
  </div>

  <!-- ATTENDANCE -->
  <div class="tab-pane fade mt-4" id="attendance" role="tabpanel" aria-labelledby="attendance-tab">
    <h3>Attendance Records</h3>
    <div class="row mb-3">
      <div class="col-md-3">
        <label class="form-label">Student ID</label>
        <input type="text" id="filter_student_id" class="form-control" placeholder="e.g. 1234" />
      </div>
      <div class="col-md-3">
        <label class="form-label">Subject ID</label>
        <input type="text" id="filter_subject_id" class="form-control" placeholder="e.g. abc123" />
      </div>
      <div class="col-md-3">
        <label class="form-label">Start Date</label>
        <input type="date" id="filter_start" class="form-control" />
      </div>
      <div class="col-md-3">
        <label class="form-label">End Date</label>
        <input type="date" id="filter_end" class="form-control" />
      </div>
    </div>
    <button class="btn btn-info mb-3" onclick="loadAttendance()">Apply Filters</button>
    <table id="attendanceTable" class="display table table-striped w-100">
      <thead>
        <tr>
          <th>Doc ID</th>
          <th>Student ID</th>
          <th>Name</th>
          <th>Subject ID</th>
          <th>Subject Name</th>
          <th>Timestamp</th>
          <th>Status</th>
        </tr>
      </thead>
      <tbody></tbody>
    </table>
    <div class="mt-3">
      <button class="btn btn-warning" onclick="saveEdits()">Save Changes</button>
      <button class="btn btn-secondary" onclick="downloadExcel()">Download Excel</button>
      <button class="btn btn-link" onclick="downloadTemplate()">Download Template</button>
      <label class="form-label d-block mt-3">Upload Excel (template must match columns):</label>
      <input type="file" id="excelFile" accept=".xlsx" class="form-control mb-2" />
      <button class="btn btn-dark" onclick="uploadExcel()">Upload Excel</button>
    </div>
  </div>
</div>

<!-- Chatbot Toggle Button -->
<button id="chatbotToggle">ðŸ’¬</button>

<!-- Chatbot Window -->
<div id="chatbotWindow" style="display:none; flex-direction:column;">
  <div id="chatHeader">
    <span>Gemini Chat</span>
    <button class="close-btn" id="chatCloseBtn">X</button>
  </div>
  <div id="chatMessages" style="flex:1; overflow-y:auto; padding:10px; font-size:14px;"></div>
  <div id="chatInputArea" style="display:flex; border-top:1px solid #ddd;">
    <input type="text" id="chatInput" placeholder="Type a message..." style="flex:1; padding:8px; border:none; outline:none; font-size:14px;" />
    <button id="chatSendBtn" style="background-color:#0d6efd; color:#fff; border:none; padding:0 15px; cursor:pointer;">Send</button>
  </div>
</div>

<!-- Scripts -->
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://code.jquery.com/jquery-3.6.0.min.js"></script>
<script src="https://cdn.datatables.net/1.13.4/js/jquery.dataTables.min.js"></script>

<script>
  /* -------------- Chatbot Code -------------- */
  const toggleBtn = document.getElementById('chatbotToggle');
  const chatWindow = document.getElementById('chatbotWindow');
  const chatCloseBtn = document.getElementById('chatCloseBtn');
  const chatMessages = document.getElementById('chatMessages');
  const chatInput = document.getElementById('chatInput');
  const chatSendBtn = document.getElementById('chatSendBtn');

  // Toggle chat window
  toggleBtn.addEventListener('click', () => {
    if (chatWindow.style.display === 'none' || chatWindow.style.display === '') {
      chatWindow.style.display = 'flex';
    } else {
      chatWindow.style.display = 'none';
    }
  });

  // Close chat window
  chatCloseBtn.addEventListener('click', () => {
    chatWindow.style.display = 'none';
  });

  function sendMessage() {
    const userMessage = chatInput.value.trim();
    if (!userMessage) return;
    addMessage(userMessage, 'user');
    chatInput.value = '';

    fetch('/process_prompt', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ prompt: userMessage })
    })
    .then(res => res.json())
    .then(data => {
      if (data.error) {
        addMessage("Error: " + data.error, 'assistant');
      } else {
        addMessage(data.message, 'assistant');
      }
    })
    .catch(err => {
      addMessage("Network or server error!", 'assistant');
      console.error(err);
    });
  }

  function addMessage(text, sender) {
    const div = document.createElement('div');
    div.classList.add('message', sender);
    div.textContent = text;
    chatMessages.appendChild(div);
    chatMessages.scrollTop = chatMessages.scrollHeight;
  }

  chatSendBtn.addEventListener('click', sendMessage);
  chatInput.addEventListener('keydown', (e) => {
    if (e.key === 'Enter') {
      e.preventDefault();
      sendMessage();
    }
  });

  /* -------------- Register + Recognize + Subjects + Attendance -------------- */
  let table;
  let attendanceData = [];

  function getBase64(file, callback) {
    const reader = new FileReader();
    reader.readAsDataURL(file);
    reader.onload = () => callback(reader.result);
    reader.onerror = (error) => console.error('Error:', error);
  }

  /* Register Face */
  function registerFace() {
    const name = document.getElementById('reg_name').value.trim();
    const studentId = document.getElementById('reg_student_id').value.trim();
    const file = document.getElementById('reg_image').files[0];
    if (!name || !studentId || !file) {
      alert('Please provide name, student ID, and an image.');
      return;
    }
    getBase64(file, (base64Str) => {
      fetch('/register', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ name, student_id: studentId, image: base64Str })
      })
      .then(res => res.json())
      .then(data => {
        const div = document.getElementById('register_result');
        div.style.display = 'block';
        div.textContent = data.message || data.error || JSON.stringify(data);
      })
      .catch(err => console.error(err));
    });
  }

  /* Recognize Faces */
  function recognizeFace() {
    const file = document.getElementById('rec_image').files[0];
    const subjectId = document.getElementById('rec_subject_select').value;
    if (!file) {
      alert('Please select an image to recognize.');
      return;
    }
    getBase64(file, (base64Str) => {
      fetch('/recognize', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ image: base64Str, subject_id: subjectId })
      })
      .then(res => res.json())
      .then(data => {
        const div = document.getElementById('recognize_result');
        div.style.display = 'block';
        let text = data.message || data.error || JSON.stringify(data);
        if (data.identified_people) {
          text += "\\n\\nIdentified People:\\n";
          data.identified_people.forEach((p) => {
            text += `- ${p.name || "Unknown"} (ID: ${p.student_id || "N/A"}), Confidence: ${p.confidence}\\n`;
          });
        }
        div.textContent = text;
      })
      .catch(err => console.error(err));
    });
  }

  /* Subjects */
  function addSubject() {
    const subjectName = document.getElementById('subject_name').value.trim();
    if (!subjectName) {
      alert('Please enter subject name.');
      return;
    }
    fetch('/add_subject', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ subject_name: subjectName })
    })
    .then(res => res.json())
    .then(data => {
      const div = document.getElementById('subject_result');
      div.style.display = 'block';
      div.textContent = data.message || data.error || JSON.stringify(data);
      document.getElementById('subject_name').value = '';
      loadSubjects();
    })
    .catch(err => console.error(err));
  }

  function loadSubjects() {
    fetch('/get_subjects')
    .then(res => res.json())
    .then(data => {
      const select = document.getElementById('rec_subject_select');
      select.innerHTML = '<option value="">-- No Subject --</option>';
      (data.subjects || []).forEach(sub => {
        const option = document.createElement('option');
        option.value = sub.id;
        option.textContent = sub.name;
        select.appendChild(option);
      });
      const list = document.getElementById('subjects_list');
      list.innerHTML = '';
      (data.subjects || []).forEach(sub => {
        const li = document.createElement('li');
        li.textContent = `ID: ${sub.id}, Name: ${sub.name}`;
        list.appendChild(li);
      });
    })
    .catch(err => console.error(err));
  }

  /* Attendance */
  function loadAttendance() {
    const studentId = document.getElementById('filter_student_id').value.trim();
    const subjectId = document.getElementById('filter_subject_id').value.trim();
    const startDate = document.getElementById('filter_start').value;
    const endDate = document.getElementById('filter_end').value;

    let url = '/api/attendance?';
    if (studentId) url += 'student_id=' + studentId + '&';
    if (subjectId) url += 'subject_id=' + subjectId + '&';
    if (startDate) url += 'start_date=' + startDate + '&';
    if (endDate) url += 'end_date=' + endDate + '&';

    fetch(url)
      .then(res => res.json())
      .then(data => {
        attendanceData = data;
        renderAttendanceTable(attendanceData);
      })
      .catch(err => console.error(err));
  }

  function renderAttendanceTable(data) {
    if ($.fn.DataTable.isDataTable('#attendanceTable')) {
      $('#attendanceTable').DataTable().clear().destroy();
    }
    const tbody = document.querySelector('#attendanceTable tbody');
    tbody.innerHTML = '';
    data.forEach(record => {
      const row = document.createElement('tr');
      row.innerHTML = `
        <td>${record.doc_id || ''}</td>
        <td contenteditable="true">${record.student_id || ''}</td>
        <td contenteditable="true">${record.name || ''}</td>
        <td contenteditable="true">${record.subject_id || ''}</td>
        <td contenteditable="true">${record.subject_name || ''}</td>
        <td contenteditable="true">${record.timestamp || ''}</td>
        <td contenteditable="true">${record.status || ''}</td>
      `;
      tbody.appendChild(row);
    });
    $('#attendanceTable').DataTable({
      paging: true,
      searching: false,
      info: false,
      responsive: true
    });
  }

  function saveEdits() {
    const rows = document.querySelectorAll('#attendanceTable tbody tr');
    const updatedRecords = [];
    rows.forEach(row => {
      const cells = row.querySelectorAll('td');
      const doc_id = cells[0].textContent.trim();
      const student_id = cells[1].textContent.trim();
      const name = cells[2].textContent.trim();
      const subject_id = cells[3].textContent.trim();
      const subject_name = cells[4].textContent.trim();
      const timestamp = cells[5].textContent.trim();
      const status = cells[6].textContent.trim();
      updatedRecords.push({ doc_id, student_id, name, subject_id, subject_name, timestamp, status });
    });
    fetch('/api/attendance/update', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ records: updatedRecords })
    })
    .then(res => res.json())
    .then(resp => {
      alert(resp.message || JSON.stringify(resp));
    })
    .catch(err => console.error(err));
  }

  function downloadExcel() {
    const studentId = document.getElementById('filter_student_id').value.trim();
    const subjectId = document.getElementById('filter_subject_id').value.trim();
    const startDate = document.getElementById('filter_start').value;
    const endDate = document.getElementById('filter_end').value;

    let url = '/api/attendance/download?';
    if (studentId) url += 'student_id=' + studentId + '&';
    if (subjectId) url += 'subject_id=' + subjectId + '&';
    if (startDate) url += 'start_date=' + startDate + '&';
    if (endDate) url += 'end_date=' + endDate + '&';

    window.location.href = url;
  }

  function downloadTemplate() {
    window.location.href = '/api/attendance/template';
  }

  function uploadExcel() {
    const fileInput = document.getElementById('excelFile');
    if (!fileInput.files.length) {
      alert('Please select an Excel file (.xlsx)');
      return;
    }
    const file = fileInput.files[0];
    const formData = new FormData();
    formData.append('file', file);

    fetch('/api/attendance/upload', {
      method: 'POST',
      body: formData
    })
    .then(res => res.json())
    .then(resp => {
      alert(resp.message || resp.error || 'Excel uploaded');
      loadAttendance();
    })
    .catch(err => console.error(err));
  }

  document.addEventListener('DOMContentLoaded', () => {
    loadSubjects();
  });
</script>
</body>
</html>
"""

# -----------------------------
# 7) Routes
# -----------------------------

# Root route to avoid "URL not found" on /
@app.route("/", methods=["GET"])
def index():
    # Return the single-page UI
    return render_template_string(INDEX_HTML)

# Register Face (GET/POST)
@app.route("/register", methods=["GET","POST"])
def register_face():
    if request.method == "GET":
        return "Welcome to /register. Please POST with {name, student_id, image} to register."

    data = request.json
    name = data.get('name')
    student_id = data.get('student_id')
    image = data.get('image')
    if not name or not student_id or not image:
        return jsonify({"message": "Missing name, student_id, or image"}), 400

    sanitized_name = "".join(c if c.isalnum() or c in "_-." else "_" for c in name)
    image_data = image.split(",")[1]
    image_bytes = base64.b64decode(image_data)

    # Enhance image before indexing
    pil_image = Image.open(io.BytesIO(image_bytes))
    enhanced_image = enhance_image(pil_image)
    buffered = io.BytesIO()
    enhanced_image.save(buffered, format="JPEG")
    enhanced_image_bytes = buffered.getvalue()

    external_image_id = f"{sanitized_name}_{student_id}"
    try:
        response = rekognition_client.index_faces(
            CollectionId=COLLECTION_ID,
            Image={'Bytes': enhanced_image_bytes},
            ExternalImageId=external_image_id,
            DetectionAttributes=['ALL'],
            QualityFilter='AUTO'
        )
    except Exception as e:
        return jsonify({"message": f"Failed to index face: {str(e)}"}), 500

    if not response.get('FaceRecords'):
        return jsonify({"message": "No face detected in the image"}), 400

    return jsonify({"message": f"Student {name} with ID {student_id} registered successfully!"}), 200

# Recognize Face (GET/POST)
@app.route("/recognize", methods=["GET","POST"])
def recognize_face():
    if request.method == "GET":
        return "Welcome to /recognize. Please POST with {image, subject_id(optional)} to detect faces."

    data = request.json
    image_str = data.get('image')
    subject_id = data.get('subject_id') or ""
    if not image_str:
        return jsonify({"message": "No image provided"}), 400

    # Optionally fetch subject name
    subject_name = ""
    if subject_id:
        sdoc = db.collection("subjects").document(subject_id).get()
        if sdoc.exists:
            subject_name = sdoc.to_dict().get("name", "")
        else:
            subject_name = "Unknown Subject"

    image_data = image_str.split(",")[1]
    image_bytes = base64.b64decode(image_data)

    # Enhance image before detection
    pil_image = Image.open(io.BytesIO(image_bytes))
    enhanced_image = enhance_image(pil_image)
    buffered = io.BytesIO()
    enhanced_image.save(buffered, format="JPEG")
    enhanced_image_bytes = buffered.getvalue()

    try:
        # Detect faces in the image
        detect_response = rekognition_client.detect_faces(
            Image={'Bytes': enhanced_image_bytes},
            Attributes=['ALL']
        )
    except Exception as e:
        return jsonify({"message": f"Failed to detect faces: {str(e)}"}), 500

    faces = detect_response.get('FaceDetails', [])
    face_count = len(faces)
    identified_people = []

    if face_count == 0:
        return jsonify({
            "message": "No faces detected in the image.",
            "total_faces": face_count,
            "identified_people": identified_people
        }), 200

    for idx, face in enumerate(faces):
        # Rekognition provides bounding box coordinates relative to image dimensions
        bbox = face['BoundingBox']
        pil_img = Image.open(io.BytesIO(enhanced_image_bytes))
        img_width, img_height = pil_img.size

        left = int(bbox['Left'] * img_width)
        top = int(bbox['Top'] * img_height)
        width = int(bbox['Width'] * img_width)
        height = int(bbox['Height'] * img_height)
        right = left + width
        bottom = top + height

        # Crop the face from the image
        cropped_face = pil_img.crop((left, top, right, bottom))
        buffer = io.BytesIO()
        cropped_face.save(buffer, format="JPEG")
        cropped_face_bytes = buffer.getvalue()

        try:
            # Search for the face in the collection
            search_response = rekognition_client.search_faces_by_image(
                CollectionId=COLLECTION_ID,
                Image={'Bytes': cropped_face_bytes},
                MaxFaces=1,
                FaceMatchThreshold=60
            )
        except Exception as e:
            identified_people.append({
                "message": f"Error searching face {idx+1}: {str(e)}",
                "confidence": "N/A"
            })
            continue

        matches = search_response.get('FaceMatches', [])
        if not matches:
            identified_people.append({
                "message": "Face not recognized",
                "confidence": "N/A"
            })
            continue

        match = matches[0]
        ext_id = match['Face']['ExternalImageId']
        confidence = match['Face']['Confidence']

        parts = ext_id.split("_", 1)
        if len(parts) == 2:
            rec_name, rec_id = parts
        else:
            rec_name, rec_id = ext_id, "Unknown"

        identified_people.append({
            "name": rec_name,
            "student_id": rec_id,
            "confidence": confidence
        })

        # If recognized, log attendance
        if rec_id != "Unknown":
            doc = {
                "student_id": rec_id,
                "name": rec_name,
                "timestamp": datetime.utcnow().isoformat(),
                "subject_id": subject_id,
                "subject_name": subject_name,
                "status": "PRESENT"
            }
            db.collection("attendance").add(doc)

    return jsonify({
        "message": f"{face_count} face(s) detected in the photo.",
        "total_faces": face_count,
        "identified_people": identified_people
    }), 200

# SUBJECTS
@app.route("/add_subject", methods=["POST"])
def add_subject():
    data = request.json
    subject_name = data.get("subject_name")
    if not subject_name:
        return jsonify({"error": "No subject_name provided"}), 400
    doc_ref = db.collection("subjects").document()
    doc_ref.set({
        "name": subject_name.strip(),
        "created_at": datetime.utcnow().isoformat()
    })
    return jsonify({"message": f"Subject '{subject_name}' added successfully!"}), 200

@app.route("/get_subjects", methods=["GET"])
def get_subjects():
    subs = db.collection("subjects").stream()
    subj_list = []
    for s in subs:
        d = s.to_dict()
        subj_list.append({"id": s.id, "name": d.get("name","")})
    return jsonify({"subjects": subj_list}), 200

# ATTENDANCE
import openpyxl
from openpyxl import Workbook

@app.route("/api/attendance", methods=["GET"])
def get_attendance():
    student_id = request.args.get("student_id")
    subject_id = request.args.get("subject_id")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = db.collection("attendance")
    if student_id:
        query = query.where("student_id", "==", student_id)
    if subject_id:
        query = query.where("subject_id", "==", subject_id)
    if start_date:
        try:
            dt_start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.where("timestamp", ">=", dt_start.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid start_date format. Use YYYY-MM-DD."}), 400
    if end_date:
        try:
            dt_end = datetime.strptime(end_date, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, microsecond=999999
            )
            query = query.where("timestamp", "<=", dt_end.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid end_date format. Use YYYY-MM-DD."}), 400

    results = query.stream()
    out_list = []
    for doc_ in results:
        dd = doc_.to_dict()
        dd["doc_id"] = doc_.id
        out_list.append(dd)

    return jsonify(out_list)

@app.route("/api/attendance/update", methods=["POST"])
def update_attendance():
    data = request.json
    records = data.get("records", [])
    for rec in records:
        doc_id = rec.get("doc_id")
        if not doc_id:
            continue
        ref = db.collection("attendance").document(doc_id)
        update_data = {
            "student_id": rec.get("student_id",""),
            "name": rec.get("name",""),
            "subject_id": rec.get("subject_id",""),
            "subject_name": rec.get("subject_name",""),
            "timestamp": rec.get("timestamp",""),
            "status": rec.get("status","")
        }
        ref.update(update_data)
    return jsonify({"message": "Attendance records updated successfully."})

@app.route("/api/attendance/download", methods=["GET"])
def download_attendance_excel():
    student_id = request.args.get("student_id")
    subject_id = request.args.get("subject_id")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    query = db.collection("attendance")
    if student_id:
        query = query.where("student_id", "==", student_id)
    if subject_id:
        query = query.where("subject_id", "==", subject_id)
    if start_date:
        try:
            dt_start = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.where("timestamp", ">=", dt_start.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid start_date format. Use YYYY-MM-DD."}), 400
    if end_date:
        try:
            dt_end = datetime.strptime(end_date, "%Y-%m-%d").replace(
                hour=23, minute=59, second=59, microsecond=999999
            )
            query = query.where("timestamp", "<=", dt_end.isoformat())
        except ValueError:
            return jsonify({"error": "Invalid end_date format. Use YYYY-MM-DD."}), 400

    results = query.stream()
    att_list = []
    for doc_ in results:
        dd = doc_.to_dict()
        dd["doc_id"] = doc_.id
        att_list.append(dd)

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    headers = ["doc_id", "student_id", "name", "subject_id", "subject_name", "timestamp", "status"]
    ws.append(headers)

    for record in att_list:
        row = [
            record.get("doc_id",""),
            record.get("student_id",""),
            record.get("name",""),
            record.get("subject_id",""),
            record.get("subject_name",""),
            record.get("timestamp",""),
            record.get("status","")
        ]
        ws.append(row)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="attendance.xlsx"
    )

@app.route("/api/attendance/template", methods=["GET"])
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Template"
    headers = ["doc_id","student_id","name","subject_id","subject_name","timestamp","status"]
    ws.append(headers)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="attendance_template.xlsx"
    )

@app.route("/api/attendance/upload", methods=["POST"])
def upload_attendance_excel():
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file.filename.endswith(".xlsx"):
        return jsonify({"error": "Please upload a .xlsx file"}), 400

    try:
        wb = openpyxl.load_workbook(file)
    except Exception as e:
        return jsonify({"error": f"Failed to read Excel file: {str(e)}"}), 400

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    expected = ("doc_id","student_id","name","subject_id","subject_name","timestamp","status")
    if not rows or rows[0] != expected:
        return jsonify({"error": "Incorrect template format"}), 400

    for row in rows[1:]:
        doc_id, student_id, name, subject_id, subject_name, timestamp, status = row
        if doc_id:
            doc_data = {
                "student_id": student_id or "",
                "name": name or "",
                "subject_id": subject_id or "",
                "subject_name": subject_name or "",
                "timestamp": timestamp or "",
                "status": status or ""
            }
            db.collection("attendance").document(doc_id).set(doc_data, merge=True)
        else:
            new_doc = {
                "student_id": student_id or "",
                "name": name or "",
                "subject_id": subject_id or "",
                "subject_name": subject_name or "",
                "timestamp": timestamp or "",
                "status": status or ""
            }
            db.collection("attendance").add(new_doc)

    return jsonify({"message": "Excel data imported successfully."})

# -----------------------------
# 8) Gemini Chat Endpoint
# -----------------------------
@app.route("/process_prompt", methods=["POST"])
def process_prompt():
    data = request.json
    user_prompt = data.get("prompt","").strip()
    if not user_prompt:
        return jsonify({"error":"No prompt provided"}), 400

    # Add user message
    conversation_memory.append({"role":"user","content":user_prompt})

    # Build conversation string
    conv_str = ""
    for msg in conversation_memory:
        if msg["role"] == "system":
            conv_str += f"System: {msg['content']}\n"
        elif msg["role"] == "user":
            conv_str += f"User: {msg['content']}\n"
        else:
            conv_str += f"Assistant: {msg['content']}\n"

    # Call Gemini
    try:
        response = model.generate_content(conv_str)
    except Exception as e:
        assistant_reply = f"Error generating response: {str(e)}"
    else:
        if not response.candidates:
            assistant_reply = "Hmm, I'm having trouble responding right now."
        else:
            parts = response.candidates[0].content.parts
            assistant_reply = "".join(part.text for part in parts).strip()

    # Add assistant reply
    conversation_memory.append({"role":"assistant","content":assistant_reply})

    if len(conversation_memory) > MAX_MEMORY:
        conversation_memory.pop(0)

    return jsonify({"message": assistant_reply})

# -----------------------------
# 9) Run App
# -----------------------------
if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)